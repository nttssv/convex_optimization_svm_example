#!/usr/bin/env python3
"""Build the ten-row hard-margin SVM teaching workbook and score figure."""

from __future__ import annotations

import os
from pathlib import Path

os.environ.setdefault("MPLCONFIGDIR", "/tmp/mpl-cache")

import matplotlib

matplotlib.use("Agg")

import matplotlib.pyplot as plt
import _pubstyle; _pubstyle.apply()
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, ScatterChart, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import DataPoint
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from linear_hard_margin_svm import (
    FEATURES,
    LOG_FEATURES,
    Preprocessor,
    feasibility_point,
    fit_primal_hard_margin,
)


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "data" / "cgh_pa_dataset.xlsx"
OUT_DIR = ROOT / "output" / "xlsx"
FIG_DIR = ROOT / "output" / "figures"
CSV_DIR = ROOT / "output" / "csv" / "hard_margin_10row"
WORKBOOK_PATH = OUT_DIR / "cgh_hard_margin_10row_teaching.xlsx"
FIGURE_PATH = FIG_DIR / "hard_margin_10row_scores.png"

NAVY = "17365D"
BLUE = "1F4E78"
MID_BLUE = "5B9BD5"
LIGHT_BLUE = "D9EAF7"
PALE_BLUE = "EAF3F8"
GREEN = "E2F0D9"
DARK_GREEN = "548235"
YELLOW = "FFF2CC"
RED = "F4CCCC"
DARK_RED = "9C0006"
GRAY = "E7E6E6"
LIGHT_GRAY = "F3F4F6"
WHITE = "FFFFFF"
BLACK = "1F1F1F"
ORANGE = "ED7D31"
PURPLE = "7030A0"

THIN_GRAY = Side(style="thin", color="B7B7B7")
MEDIUM_BLUE = Side(style="medium", color=BLUE)


def title(ws, text: str, end_col: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    cell = ws.cell(1, 1, text)
    cell.font = Font(name="Arial", size=16, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28


def header_row(ws, row: int, start_col: int, end_col: int) -> None:
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row, col)
        cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=MEDIUM_BLUE, bottom=MEDIUM_BLUE)
    ws.row_dimensions[row].height = 38


def section_band(ws, row: int, text: str, end_col: int) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    cell = ws.cell(row, 1, text)
    cell.font = Font(name="Arial", size=11, bold=True, color=NAVY)
    cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 22


def body_grid(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for row in ws.iter_rows(
        min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
    ):
        for cell in row:
            cell.font = Font(name="Arial", size=10, color=BLACK)
            cell.border = Border(bottom=THIN_GRAY)
            cell.alignment = Alignment(vertical="center")
        if row[0].row % 2 == 0:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)


def set_widths(ws, widths: dict[int, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def add_table(ws, ref: str, name: str) -> None:
    tab = Table(displayName=name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)


def base_font(workbook: Workbook) -> None:
    for ws in workbook.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and cell.font.name != "Arial":
                    cell.font = Font(
                        name="Arial",
                        size=cell.font.sz or 10,
                        bold=cell.font.bold,
                        italic=cell.font.italic,
                        color=cell.font.color,
                    )


def build_start_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Start_Here"
    title(ws, "Linear hard-margin SVM: 10-row worked Excel example", 8)
    ws.merge_cells("A3:H3")
    ws["A3"] = (
        "TEACHING SUBSET ONLY - the complete 114-row cohort and every CV training fold "
        "remain hard-margin infeasible; no ROC is calculated from these 10 rows."
    )
    ws["A3"].font = Font(name="Arial", size=11, bold=True, color=DARK_RED)
    ws["A3"].fill = PatternFill("solid", fgColor=YELLOW)
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[3].height = 42

    section_band(ws, 5, "The formula that caused confusion", 8)
    formulas = [
        ("Parameter vector", "theta = [w1, ..., w10, b]^T"),
        ("Signed row vector", "u_i = t_i [x_i1, ..., x_i10, 1]^T"),
        ("Row calculation", "u_i^T theta = t_i (w^T x_i + b) = t_i s_i"),
        ("Hard-margin rule", "u_i^T theta >= 1"),
    ]
    for row, (label, value) in enumerate(formulas, start=6):
        ws.cell(row, 1, label).font = Font(name="Arial", bold=True, color=NAVY)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        ws.cell(row, 2, value).font = Font(name="Arial", size=11)
    ws.merge_cells("A11:H11")
    ws["A11"] = (
        "Important: U_i is a row vector, so it is not U_i > 1 by itself. "
        "The scalar dot product u_i^T theta must be at least 1."
    )
    ws["A11"].font = Font(name="Arial", bold=True, color=DARK_RED)
    ws["A11"].fill = PatternFill("solid", fgColor=RED)
    ws["A11"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[11].height = 38

    section_band(ws, 13, "How +1 and -1 become one inequality", 8)
    cases = [
        ["UPA", "+1", "s_i >= +1", "+1 x positive score >= 1"],
        ["BPA", "-1", "s_i <= -1", "-1 x negative score >= 1"],
    ]
    for col, value in enumerate(["Class y_i", "t_i", "Required score", "Signed-margin interpretation"], 1):
        ws.cell(14, col, value)
    header_row(ws, 14, 1, 4)
    for row_idx, values in enumerate(cases, start=15):
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row_idx, col_idx, value)
    body_grid(ws, 15, 16, 1, 4)

    section_band(ws, 18, "Workbook map", 8)
    mapping = [
        ["1", "Raw_10", "The first 10 source rows: every raw x_i component, class y_i, and numeric t_i."],
        ["2", "Transform_10", "Natural-log transforms plus the mean and population SD used by this subset."],
        ["3", "X_Standardized", "The standardized model vectors x_i used in the SVM formulas."],
        ["4", "Theta_Derivatives", "Optimized w and b, objective, first derivative, and full Hessian."],
        ["5", "U_Margin_Check", "Every u_i component, u_i^T theta, g_i, feasibility, and support-vector status."],
        ["6", "Visualize", "Decision scores with -1, 0, +1 reference lines and signed-margin chart."],
        ["7", "Full_Data_Gate", "Why these 10 illustrative rows do not authorize a full-data ROC."],
    ]
    for col, value in enumerate(["Step", "Sheet", "What to inspect"], 1):
        ws.cell(19, col, value)
    header_row(ws, 19, 1, 3)
    for row_idx, values in enumerate(mapping, start=20):
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row_idx, col_idx, value)
    body_grid(ws, 20, 26, 1, 3)
    for row in range(20, 27):
        ws.cell(row, 3).alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[row].height = 30

    section_band(ws, 28, "Selection and preprocessing notes", 8)
    notes = [
        "Selection rule: scan top-to-bottom, take the first five UPA and first five BPA rows, then restore source order. This equals source rows 2-11.",
        "No selected value is missing, so median imputation is inactive in this worked example.",
        "PAC, PRA, 18-OHF, and 18-oxoF use LN(value + 1E-6); all 10 transformed features use subset mean and population SD.",
        "The coefficient cells are solver outputs. If raw inputs change, run the constrained optimizer again; Excel formulas alone do not re-optimize theta.",
    ]
    for row, note in enumerate(notes, start=29):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        ws.cell(row, 1, f"- {note}")
        ws.cell(row, 1).alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[row].height = 32

    set_widths(ws, {1: 18, 2: 24, 3: 66, 4: 23, 5: 14, 6: 14, 7: 14, 8: 14})
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False


def build_raw_sheet(wb: Workbook, subset: pd.DataFrame) -> None:
    ws = wb.create_sheet("Raw_10")
    title(ws, "Step 1 - Raw values for each x_i and class label", 15)
    ws.merge_cells("A2:O2")
    ws["A2"] = (
        "Source: data/cgh_pa_dataset.xlsx, Sheet1. Selection is deterministic: first five UPA + first five BPA, restored to source order."
    )
    ws["A2"].font = Font(name="Arial", italic=True, color="666666")
    ws["A2"].alignment = Alignment(wrap_text=True)

    headers = [
        "i",
        "Source Excel row",
        "PatientID",
        "Class y_i",
        "Numeric t_i",
        *FEATURES,
    ]
    for col, value in enumerate(headers, 1):
        ws.cell(4, col, value)
    header_row(ws, 4, 1, 15)
    ws["D4"].comment = Comment("Diagnosis label: UPA or BPA.", "OpenAI")
    ws["E4"].comment = Comment("UPA = +1 and BPA = -1 for the SVM constraint.", "OpenAI")

    for i, (_, row) in enumerate(subset.iterrows(), start=1):
        excel_row = int(row.name) + 2
        values = [
            i,
            excel_row,
            row["PatientID"],
            row["Model 1 Target"],
            1 if row["Model 1 Target"] == "UPA" else -1,
            *[float(row[feature]) for feature in FEATURES],
        ]
        for col, value in enumerate(values, 1):
            ws.cell(i + 4, col, value)
    body_grid(ws, 5, 14, 1, 15)
    for row in range(5, 15):
        ws.cell(row, 5).fill = PatternFill("solid", fgColor=YELLOW)
        for col in range(6, 16):
            ws.cell(row, col).number_format = "0.000"

    add_table(ws, "A4:O14", "Raw10Table")
    set_widths(
        ws,
        {
            1: 6,
            2: 16,
            3: 14,
            4: 12,
            5: 12,
            6: 12,
            7: 12,
            8: 12,
            9: 13,
            10: 12,
            11: 12,
            12: 13,
            13: 13,
            14: 10,
            15: 10,
        },
    )
    ws.freeze_panes = "F5"
    ws.sheet_view.showGridLines = False
    ws.auto_filter.ref = "A4:O14"


def build_transform_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Transform_10")
    title(ws, "Step 2 - Transform features and calculate subset statistics", 12)
    ws.merge_cells("A2:L2")
    ws["A2"] = (
        "Formulas are live. The four hormone measurements use LN(value + 1E-6); other columns link directly to Raw_10."
    )
    ws["A2"].font = Font(name="Arial", italic=True, color="666666")
    ws["A2"].alignment = Alignment(wrap_text=True)

    display_features = [
        "LN(PAC+1E-6)",
        "LN(PRA+1E-6)",
        "Potassium",
        "Tumor size",
        "LN(18-OHF+1E-6)",
        "LN(18-oxoF+1E-6)",
        "Systolic BP",
        "Diastolic BP",
        "DDD",
        "Age",
    ]
    for col, value in enumerate(["i", "PatientID", *display_features], 1):
        ws.cell(4, col, value)
    header_row(ws, 4, 1, 12)

    raw_cols = list(range(6, 16))
    for out_row, raw_row in zip(range(5, 15), range(5, 15)):
        ws.cell(out_row, 1, f"='Raw_10'!A{raw_row}")
        ws.cell(out_row, 2, f"='Raw_10'!C{raw_row}")
        for feature_idx, raw_col in enumerate(raw_cols, start=0):
            out_col = feature_idx + 3
            source = f"'Raw_10'!{get_column_letter(raw_col)}{raw_row}"
            feature = FEATURES[feature_idx]
            ws.cell(out_row, out_col, f"=LN({source}+1E-6)" if feature in LOG_FEATURES else f"={source}")
            ws.cell(out_row, out_col).number_format = "0.000000"
    body_grid(ws, 5, 14, 1, 12)

    ws["A16"] = "Mean (mu_j)"
    ws["A17"] = "Population SD (sigma_j)"
    ws["A18"] = "Median (imputation reference)"
    for col in range(3, 13):
        letter = get_column_letter(col)
        ws.cell(16, col, f"=AVERAGE({letter}5:{letter}14)")
        ws.cell(17, col, f"=STDEV.P({letter}5:{letter}14)")
        ws.cell(18, col, f"=MEDIAN({letter}5:{letter}14)")
        for row in (16, 17, 18):
            ws.cell(row, col).number_format = "0.000000000"
    for row in range(16, 19):
        for col in range(1, 13):
            ws.cell(row, col).fill = PatternFill("solid", fgColor=PALE_BLUE)
            ws.cell(row, col).font = Font(name="Arial", bold=(col == 1), color=BLACK)
            ws.cell(row, col).border = Border(bottom=THIN_GRAY)

    ws.merge_cells("A20:L20")
    ws["A20"] = (
        "No values are missing in these ten rows, so the median is shown for completeness but is not substituted anywhere."
    )
    ws["A20"].font = Font(name="Arial", italic=True, color="666666")
    set_widths(ws, {1: 6, 2: 14, **{col: 16 for col in range(3, 13)}})
    ws.freeze_panes = "C5"
    ws.sheet_view.showGridLines = False


def build_standardized_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("X_Standardized")
    title(ws, "Step 3 - Standardized model vectors x_i", 14)
    ws.merge_cells("A2:N2")
    ws["A2"] = "Each component is x_ij = (transformed value - mu_j) / sigma_j. These x_i vectors enter the SVM."
    ws["A2"].font = Font(name="Arial", italic=True, color="666666")

    for col, value in enumerate(["i", "PatientID", "Class y_i", "t_i", *FEATURES], 1):
        ws.cell(4, col, value)
    header_row(ws, 4, 1, 14)

    for out_row, source_row in zip(range(5, 15), range(5, 15)):
        ws.cell(out_row, 1, f"='Raw_10'!A{source_row}")
        ws.cell(out_row, 2, f"='Raw_10'!C{source_row}")
        ws.cell(out_row, 3, f"='Raw_10'!D{source_row}")
        ws.cell(out_row, 4, f"='Raw_10'!E{source_row}")
        for feature_idx in range(10):
            out_col = feature_idx + 5
            transform_col = feature_idx + 3
            letter = get_column_letter(transform_col)
            ws.cell(
                out_row,
                out_col,
                f"=('Transform_10'!{letter}{source_row}-'Transform_10'!{letter}$16)/'Transform_10'!{letter}$17",
            )
            ws.cell(out_row, out_col).number_format = "0.000000000"
    body_grid(ws, 5, 14, 1, 14)
    for row in range(5, 15):
        ws.cell(row, 4).fill = PatternFill("solid", fgColor=YELLOW)

    ws["A16"] = "Check mean"
    ws["A17"] = "Check population SD"
    for col in range(5, 15):
        letter = get_column_letter(col)
        ws.cell(16, col, f"=AVERAGE({letter}5:{letter}14)")
        ws.cell(17, col, f"=STDEV.P({letter}5:{letter}14)")
        ws.cell(16, col).number_format = "0.000000000"
        ws.cell(17, col).number_format = "0.000000000"
    for row in (16, 17):
        for col in range(1, 15):
            ws.cell(row, col).fill = PatternFill("solid", fgColor=PALE_BLUE)
            ws.cell(row, col).font = Font(name="Arial", bold=(col == 1), color=BLACK)
            ws.cell(row, col).border = Border(bottom=THIN_GRAY)

    ws.merge_cells("A19:N19")
    ws["A19"] = "Expected checks: feature means approximately 0 and population SDs approximately 1."
    ws["A19"].font = Font(name="Arial", italic=True, color="666666")
    set_widths(ws, {1: 6, 2: 14, 3: 12, 4: 9, **{col: 15 for col in range(5, 15)}})
    ws.freeze_panes = "E5"
    ws.sheet_view.showGridLines = False


def build_theta_sheet(wb: Workbook, theta: np.ndarray) -> None:
    ws = wb.create_sheet("Theta_Derivatives")
    title(ws, "Step 4 - Values optimized in the primal, gradient, and Hessian", 13)
    ws.merge_cells("A2:M2")
    ws["A2"] = (
        "Optimize theta = [w1,...,w10,b]^T to minimize 0.5||w||^2 subject to every u_i^T theta >= 1. "
        "Blue numbers are solver outputs; black cells are Excel formulas."
    )
    ws["A2"].font = Font(name="Arial", italic=True, color="666666")
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[2].height = 35

    headers = [
        "Index",
        "theta component",
        "Feature / role",
        "Optimized value",
        "0.5*w_j^2 contribution",
        "First derivative d(phi)/d(theta_j)",
        "Hessian diagonal",
    ]
    for col, value in enumerate(headers, 1):
        ws.cell(4, col, value)
    header_row(ws, 4, 1, 7)

    roles = FEATURES + ["Intercept b"]
    for idx in range(11):
        row = idx + 5
        ws.cell(row, 1, idx + 1)
        ws.cell(row, 2, f"w{idx + 1}" if idx < 10 else "b")
        ws.cell(row, 3, roles[idx])
        ws.cell(row, 4, float(theta[idx]))
        ws.cell(row, 4).number_format = "0.000000000"
        ws.cell(row, 4).font = Font(name="Arial", color="0000FF")
        ws.cell(row, 4).comment = Comment(
            "Hard-margin constrained-QP solver output. Re-optimize if the raw inputs change.",
            "OpenAI",
        )
        ws.cell(row, 5, f"=0.5*D{row}^2" if idx < 10 else "=0")
        ws.cell(row, 6, f"=D{row}" if idx < 10 else "=0")
        ws.cell(row, 7, 1 if idx < 10 else 0)
        for col in range(5, 8):
            ws.cell(row, col).number_format = "0.000000000"
    body_grid(ws, 5, 15, 1, 7)
    for row in range(5, 16):
        ws.cell(row, 4).font = Font(name="Arial", color="0000FF")

    section_band(ws, 18, "Objective and margin summary", 7)
    summary = [
        ("Primal objective phi(theta) = 0.5||w||^2", "=SUM(E5:E15)", "Minimized value"),
        ("||w||", "=SQRT(SUMSQ(D5:D14))", "Normal-vector length"),
        ("Boundary-to-support distance = 1/||w||", "=1/B20", "One side of margin"),
        ("Full margin width = 2/||w||", "=2/B20", "Distance between score -1 and +1"),
        ("Intercept b", "=D15", "Not penalized in the objective"),
        ("10-row solver status", "Optimal and feasible", "Teaching subsystem only"),
    ]
    for offset, (label, value, meaning) in enumerate(summary, start=19):
        ws.cell(offset, 1, label)
        ws.cell(offset, 2, value)
        ws.cell(offset, 3, meaning)
        ws.cell(offset, 1).font = Font(name="Arial", bold=True, color=NAVY)
        ws.cell(offset, 2).number_format = "0.000000000"
        ws.cell(offset, 3).font = Font(name="Arial", italic=True, color="666666")
    for row in range(19, 25):
        for col in range(1, 4):
            ws.cell(row, col).border = Border(bottom=THIN_GRAY)

    section_band(ws, 27, "Full second-derivative matrix Q = diag(I_10, 0)", 13)
    components = [f"w{i}" for i in range(1, 11)] + ["b"]
    ws.cell(28, 1, "d2(phi)")
    for idx, comp in enumerate(components, start=2):
        ws.cell(28, idx, comp)
        ws.cell(idx + 27, 1, comp)
    header_row(ws, 28, 1, 12)
    for r_idx in range(11):
        row = r_idx + 29
        ws.cell(row, 1).font = Font(name="Arial", bold=True, color=NAVY)
        ws.cell(row, 1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        for c_idx in range(11):
            col = c_idx + 2
            value = 1 if r_idx == c_idx and r_idx < 10 else 0
            ws.cell(row, col, value)
            ws.cell(row, col).alignment = Alignment(horizontal="center")
            ws.cell(row, col).border = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)
            ws.cell(row, col).fill = PatternFill("solid", fgColor=GREEN if value == 1 else LIGHT_GRAY)
    ws.merge_cells("A41:M41")
    ws["A41"] = (
        "The intercept row and column are zero because b is not part of 0.5||w||^2. "
        "For every affine constraint g_i=1-u_i^T theta, the entire second derivative is zero."
    )
    ws["A41"].font = Font(name="Arial", italic=True, color="666666")
    ws["A41"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[41].height = 34

    set_widths(ws, {1: 36, 2: 18, 3: 26, 4: 18, 5: 22, 6: 27, 7: 18, **{col: 9 for col in range(8, 14)}})
    ws.freeze_panes = "D5"
    ws.sheet_view.showGridLines = False


def build_margin_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("U_Margin_Check")
    title(ws, "Step 5 - Row-by-row U theta >= 1 calculation", 24)
    ws.merge_cells("A2:X2")
    ws["A2"] = (
        "For row i: u_i = t_i[x_i,1]. The scalar u_i^T theta equals t_i*s_i and must be >= 1. "
        "This is the correct interpretation, not 'U_i > 1'."
    )
    ws["A2"].font = Font(name="Arial", bold=True, color=DARK_RED)
    ws["A2"].fill = PatternFill("solid", fgColor=YELLOW)
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[2].height = 38
    ws.merge_cells("A3:X3")
    ws["A3"] = (
        "Constraint function g_i(theta)=1-u_i^T theta must be <= 0. Its first derivative is -u_i; its second derivative is the zero matrix."
    )
    ws["A3"].font = Font(name="Arial", italic=True, color="666666")

    headers = [
        "i",
        "PatientID",
        "Class y_i",
        "t_i",
        *[f"u_{feature}" for feature in FEATURES],
        "u_b=t_i",
        "Score s_i",
        "u_i^T theta",
        "t_i*s_i",
        "Identity",
        "g_i=1-u_i^T theta",
        "Constraint",
        "Support vector",
        "Predicted t",
        "Correct?",
    ]
    for col, value in enumerate(headers, 1):
        ws.cell(5, col, value)
    header_row(ws, 5, 1, 24)

    for row_idx in range(10):
        row = row_idx + 6
        source_row = row_idx + 5
        ws.cell(row, 1, f"='X_Standardized'!A{source_row}")
        ws.cell(row, 2, f"='X_Standardized'!B{source_row}")
        ws.cell(row, 3, f"='X_Standardized'!C{source_row}")
        ws.cell(row, 4, f"='X_Standardized'!D{source_row}")
        for feature_idx in range(10):
            u_col = feature_idx + 5
            x_col = get_column_letter(feature_idx + 5)
            ws.cell(row, u_col, f"=$D{row}*'X_Standardized'!{x_col}{source_row}")
            ws.cell(row, u_col).number_format = "0.000000"
        ws.cell(row, 15, f"=$D{row}")
        ws.cell(
            row,
            16,
            f"=SUMPRODUCT('X_Standardized'!E{source_row}:N{source_row},'Theta_Derivatives'!$D$5:$D$14)+'Theta_Derivatives'!$D$15",
        )
        ws.cell(row, 17, f"=SUMPRODUCT(E{row}:O{row},'Theta_Derivatives'!$D$5:$D$15)")
        ws.cell(row, 18, f"=$D{row}*$P{row}")
        ws.cell(row, 19, f'=IF(ABS($Q{row}-$R{row})<=1E-8,"OK","ERROR")')
        ws.cell(row, 20, f"=1-$Q{row}")
        ws.cell(row, 21, f'=IF($T{row}<=1E-8,"Satisfied","Violated")')
        ws.cell(row, 22, f'=IF(ABS($T{row})<=1E-5,"Yes","No")')
        ws.cell(row, 23, f'=IF($P{row}>=0,1,-1)')
        ws.cell(row, 24, f'=IF($W{row}=$D{row},"Correct","Wrong")')
        for col in (16, 17, 18, 20):
            ws.cell(row, col).number_format = "0.000000000"
    body_grid(ws, 6, 15, 1, 24)
    for row in range(6, 16):
        ws.cell(row, 4).fill = PatternFill("solid", fgColor=YELLOW)
        ws.cell(row, 17).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        ws.cell(row, 21).fill = PatternFill("solid", fgColor=GREEN)

    ws.conditional_formatting.add(
        "U6:U15",
        FormulaRule(formula=['U6="Violated"'], fill=PatternFill("solid", fgColor=RED), font=Font(color=DARK_RED)),
    )
    ws.conditional_formatting.add(
        "S6:S15",
        FormulaRule(formula=['S6="ERROR"'], fill=PatternFill("solid", fgColor=RED), font=Font(color=DARK_RED)),
    )
    ws.conditional_formatting.add(
        "X6:X15",
        FormulaRule(formula=['X6="Wrong"'], fill=PatternFill("solid", fgColor=RED), font=Font(color=DARK_RED)),
    )

    section_band(ws, 18, "Constraint summary", 6)
    summary = [
        ("Minimum signed margin", "=MIN(Q6:Q15)"),
        ("All ten constraints satisfied?", '=IF(MIN(Q6:Q15)>=1-1E-8,"YES","NO")'),
        ("Support-vector count", '=COUNTIF(V6:V15,"Yes")'),
        ("Training accuracy (illustrative only)", '=COUNTIF(X6:X15,"Correct")/ROWS(X6:X15)'),
    ]
    for row, (label, formula) in enumerate(summary, start=19):
        ws.cell(row, 1, label)
        ws.cell(row, 2, formula)
        ws.cell(row, 1).font = Font(name="Arial", bold=True, color=NAVY)
        ws.cell(row, 1).border = Border(bottom=THIN_GRAY)
        ws.cell(row, 2).border = Border(bottom=THIN_GRAY)
    ws["B19"].number_format = "0.000000000"
    ws["B22"].number_format = "0.0%"
    ws.merge_cells("A24:X24")
    ws["A24"] = (
        "Support vectors have u_i^T theta approximately equal to 1. Correct classification alone needs t_i*s_i > 0; hard margin is stricter and needs t_i*s_i >= 1."
    )
    ws["A24"].font = Font(name="Arial", italic=True, color="666666")
    ws["A24"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[24].height = 32

    widths = {1: 6, 2: 14, 3: 12, 4: 8}
    widths.update({col: 15 for col in range(5, 16)})
    widths.update({16: 14, 17: 15, 18: 14, 19: 12, 20: 18, 21: 14, 22: 15, 23: 12, 24: 12})
    set_widths(ws, widths)
    ws.freeze_panes = "E6"
    ws.sheet_view.showGridLines = False
    ws.auto_filter.ref = "A5:X15"


def build_visual_sheet(wb: Workbook, subset: pd.DataFrame) -> None:
    ws = wb.create_sheet("Visualize")
    title(ws, "Step 6 - Visualize scores, margins, and the decision boundary", 18)
    ws.merge_cells("A2:R2")
    ws["A2"] = (
        "Raw decision scores use the -1, 0, +1 reference lines. Signed margins place both classes on the positive side and are checked against 1."
    )
    ws["A2"].font = Font(name="Arial", italic=True, color="666666")

    headers = ["i", "PatientID", "Class", "t_i", "Score s_i", "Signed margin", "Target margin"]
    for col, value in enumerate(headers, 1):
        ws.cell(5, col, value)
    header_row(ws, 5, 1, 7)
    for idx in range(10):
        row = idx + 6
        source_row = idx + 6
        ws.cell(row, 1, f"='U_Margin_Check'!A{source_row}")
        ws.cell(row, 2, f"='U_Margin_Check'!B{source_row}")
        ws.cell(row, 3, f"='U_Margin_Check'!C{source_row}")
        ws.cell(row, 4, f"='U_Margin_Check'!D{source_row}")
        ws.cell(row, 5, f"='U_Margin_Check'!P{source_row}")
        ws.cell(row, 6, f"='U_Margin_Check'!Q{source_row}")
        ws.cell(row, 7, 1)
        ws.cell(row, 5).number_format = "0.000"
        ws.cell(row, 6).number_format = "0.000"
    body_grid(ws, 6, 15, 1, 7)

    ws["I5"] = "UPA score"
    ws["J5"] = "Observation i"
    ws["L5"] = "BPA score"
    ws["M5"] = "Observation i"
    for col in (9, 10, 12, 13):
        ws.cell(5, col).font = Font(name="Arial", bold=True, color=WHITE)
        ws.cell(5, col).fill = PatternFill("solid", fgColor=BLUE)

    upa_rows = [idx + 6 for idx, value in enumerate(subset["Model 1 Target"]) if value == "UPA"]
    bpa_rows = [idx + 6 for idx, value in enumerate(subset["Model 1 Target"]) if value == "BPA"]
    for out_row, source_row in zip(range(6, 11), upa_rows):
        ws.cell(out_row, 9, f"=E{source_row}")
        ws.cell(out_row, 10, f"=A{source_row}")
    for out_row, source_row in zip(range(6, 11), bpa_rows):
        ws.cell(out_row, 12, f"=E{source_row}")
        ws.cell(out_row, 13, f"=A{source_row}")

    refs = [(15, -1, "Margin -1"), (17, 0, "Boundary 0"), (19, 1, "Margin +1")]
    for start_col, x_value, label in refs:
        ws.cell(5, start_col, label)
        ws.cell(5, start_col + 1, "Observation i")
        ws.cell(6, start_col, x_value)
        ws.cell(7, start_col, x_value)
        ws.cell(6, start_col + 1, 0.5)
        ws.cell(7, start_col + 1, 10.5)

    scatter = ScatterChart()
    scatter.title = "Illustrative hard-margin scores (10-row teaching subset)"
    scatter.style = 13
    scatter.height = 12
    scatter.width = 23
    scatter.x_axis.title = "Decision score s_i = w^T x_i + b"
    scatter.y_axis.title = "Observation i"
    scatter.x_axis.scaling.min = -3.5
    scatter.x_axis.scaling.max = 3.5
    scatter.y_axis.scaling.min = 0.5
    scatter.y_axis.scaling.max = 10.5
    scatter.y_axis.majorUnit = 1
    scatter.legend.position = "b"

    upa_series = Series(Reference(ws, min_col=9, min_row=6, max_row=10), Reference(ws, min_col=10, min_row=6, max_row=10), title="UPA (t=+1)")
    upa_series.marker.symbol = "circle"
    upa_series.marker.size = 9
    upa_series.graphicalProperties.line.noFill = True
    upa_series.graphicalProperties.solidFill = MID_BLUE
    bpa_series = Series(Reference(ws, min_col=12, min_row=6, max_row=10), Reference(ws, min_col=13, min_row=6, max_row=10), title="BPA (t=-1)")
    bpa_series.marker.symbol = "diamond"
    bpa_series.marker.size = 9
    bpa_series.graphicalProperties.line.noFill = True
    bpa_series.graphicalProperties.solidFill = ORANGE
    scatter.series.append(upa_series)
    scatter.series.append(bpa_series)

    for start_col, _, label in refs:
        ref_series = Series(
            Reference(ws, min_col=start_col, min_row=6, max_row=7),
            Reference(ws, min_col=start_col + 1, min_row=6, max_row=7),
            title=label,
        )
        ref_series.marker.symbol = "none"
        ref_series.graphicalProperties.line.solidFill = BLACK if label == "Boundary 0" else "808080"
        ref_series.graphicalProperties.line.width = 22000 if label == "Boundary 0" else 13000
        if label != "Boundary 0":
            ref_series.graphicalProperties.line.prstDash = "dash"
        scatter.series.append(ref_series)
    ws.add_chart(scatter, "A18")

    line = LineChart()
    line.title = "Signed margin u_i^T theta compared with target 1"
    line.style = 10
    line.height = 11
    line.width = 22
    line.y_axis.title = "Signed functional margin"
    line.x_axis.title = "Patient"
    line.y_axis.scaling.min = 0
    line.y_axis.scaling.max = 3.5
    line.y_axis.majorGridlines = None
    data = Reference(ws, min_col=6, max_col=7, min_row=5, max_row=15)
    categories = Reference(ws, min_col=2, min_row=6, max_row=15)
    line.add_data(data, titles_from_data=True)
    line.set_categories(categories)
    line.legend.position = "b"
    line.series[0].graphicalProperties.line.solidFill = MID_BLUE
    line.series[0].marker.symbol = "circle"
    line.series[0].marker.size = 7
    line.series[1].graphicalProperties.line.solidFill = ORANGE
    line.series[1].graphicalProperties.line.prstDash = "dash"
    ws.add_chart(line, "J18")

    ws.merge_cells("A42:R42")
    ws["A42"] = (
        "Interpretation: UPA must lie at score >= +1, BPA must lie at score <= -1, and score 0 is the classifier boundary. Scores are not probabilities."
    )
    ws["A42"].font = Font(name="Arial", bold=True, color=NAVY)
    ws["A42"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    ws["A42"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[42].height = 34
    set_widths(ws, {1: 7, 2: 14, 3: 11, 4: 8, 5: 14, 6: 16, 7: 15, 9: 13, 10: 13, 12: 13, 13: 13, 15: 13, 16: 13, 17: 13, 18: 13})
    ws.freeze_panes = "A6"
    ws.sheet_view.showGridLines = False


def build_full_data_gate(wb: Workbook) -> None:
    ws = wb.create_sheet("Full_Data_Gate")
    title(ws, "Step 7 - Full-data feasibility gate and ROC decision", 7)
    ws.merge_cells("A3:G3")
    ws["A3"] = (
        "The 10-row worksheet is feasible because it omits 104 constraints. Adding the remaining observations makes the full linear hard-margin system infeasible."
    )
    ws["A3"].font = Font(name="Arial", bold=True, color=DARK_RED)
    ws["A3"].fill = PatternFill("solid", fgColor=YELLOW)
    ws["A3"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[3].height = 42

    headers = ["Analysis instance", "Training n", "Held-out n", "Hard-margin feasible?", "Model fitted?", "ROC scores?", "AUROC"]
    for col, value in enumerate(headers, 1):
        ws.cell(5, col, value)
    header_row(ws, 5, 1, 7)
    rows = [
        ["Full cohort", 114, "-", "No", "No", "No", "N/A"],
        ["CV fold 1", 91, 23, "No", "No", "No", "N/A"],
        ["CV fold 2", 91, 23, "No", "No", "No", "N/A"],
        ["CV fold 3", 91, 23, "No", "No", "No", "N/A"],
        ["CV fold 4", 91, 23, "No", "No", "No", "N/A"],
        ["CV fold 5", 92, 22, "No", "No", "No", "N/A"],
    ]
    for row_idx, values in enumerate(rows, start=6):
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row_idx, col_idx, value)
    body_grid(ws, 6, 11, 1, 7)
    for row in range(6, 12):
        for col in range(4, 8):
            ws.cell(row, col).fill = PatternFill("solid", fgColor=RED)
            ws.cell(row, col).font = Font(name="Arial", bold=True, color=DARK_RED)

    section_band(ws, 14, "ROC rule", 7)
    ws.merge_cells("A15:G16")
    ws["A15"] = (
        "A ROC curve needs a valid fitted scoring function for every held-out observation. Because each training fold is infeasible, there are no valid out-of-fold scores. "
        "Therefore ROC, AUROC, sensitivity, and specificity are not applicable for the requested full-data linear hard-margin model. The 10-row in-sample separation is not a performance estimate."
    )
    ws["A15"].alignment = Alignment(wrap_text=True, vertical="center")
    ws["A15"].font = Font(name="Arial", size=11)
    ws["A15"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    set_widths(ws, {1: 24, 2: 14, 3: 14, 4: 22, 5: 16, 6: 15, 7: 12})
    ws.freeze_panes = "A6"
    ws.sheet_view.showGridLines = False


def save_score_figure(subset: pd.DataFrame, scores: np.ndarray) -> None:
    labels = np.where(subset["Model 1 Target"].eq("UPA"), 1, -1)
    y = np.arange(1, len(subset) + 1)
    fig, ax = plt.subplots(figsize=(8.2, 5.2), constrained_layout=True)
    for class_value, color, marker, name in [
        (1, "#2f75b5", "o", "UPA ($t_i=+1$)"),
        (-1, "#ed7d31", "D", "BPA ($t_i=-1$)"),
    ]:
        mask = labels == class_value
        ax.scatter(scores[mask], y[mask], s=70, color=color, marker=marker, label=name, zorder=3)
    ax.axvline(-1, color="#808080", linestyle="--", linewidth=1.5, label="Margins $s=-1,+1$")
    ax.axvline(0, color="black", linewidth=1.8, label="Decision boundary $s=0$")
    ax.axvline(1, color="#808080", linestyle="--", linewidth=1.5)
    ax.set(
        xlabel=r"Decision score $s_i=w^\top x_i+b$",
        ylabel="Observation",
        title="Illustrative linear hard-margin scores for the 10-row teaching subset",
        xlim=(-3.5, 3.5),
        ylim=(10.7, 0.3),
    )
    ax.set_yticks(y, subset["PatientID"].astype(str))
    ax.grid(axis="x", alpha=0.2)
    ax.legend(loc="lower right", frameon=True)
    fig.savefig(FIGURE_PATH, dpi=300, facecolor="white")
    plt.close(fig)


def export_csv_tables(
    selected: pd.DataFrame,
    X_raw: pd.DataFrame,
    preprocessor: Preprocessor,
    X: np.ndarray,
    labels: np.ndarray,
    theta: np.ndarray,
    scores: np.ndarray,
) -> None:
    CSV_DIR.mkdir(parents=True, exist_ok=True)
    identifiers = pd.DataFrame(
        {
            "i": np.arange(1, 11),
            "source_excel_row": selected.index.to_numpy() + 2,
            "PatientID": selected["PatientID"].to_numpy(),
            "class_y_i": selected["Model 1 Target"].to_numpy(),
            "t_i": labels.astype(int),
        }
    )
    raw = pd.concat(
        [identifiers.reset_index(drop=True), X_raw.reset_index(drop=True)], axis=1
    )
    raw.to_csv(CSV_DIR / "01_raw_x_and_y.csv", index=False)

    transformed = X_raw.copy()
    for feature in LOG_FEATURES:
        transformed[feature] = np.log(transformed[feature] + 1e-6)
    transforms = ["LN(value + 1E-6)" if f in LOG_FEATURES else "unchanged" for f in FEATURES]
    pd.DataFrame(
        {
            "feature": FEATURES,
            "transform": transforms,
            "mean_mu_j": preprocessor.means.reindex(FEATURES).to_numpy(),
            "population_sd_sigma_j": preprocessor.scales.reindex(FEATURES).to_numpy(),
            "median_imputation_reference": preprocessor.medians.reindex(FEATURES).to_numpy(),
        }
    ).to_csv(CSV_DIR / "02_preprocessing_parameters.csv", index=False)
    transformed_out = pd.concat(
        [identifiers.iloc[:, [0, 2]].reset_index(drop=True), transformed.reset_index(drop=True)], axis=1
    )
    transformed_out.to_csv(CSV_DIR / "03_transformed_x.csv", index=False)

    standardized = pd.DataFrame(X, columns=[f"x_{feature}" for feature in FEATURES])
    standardized_out = pd.concat(
        [identifiers.reset_index(drop=True), standardized], axis=1
    )
    standardized_out.to_csv(CSV_DIR / "04_standardized_x.csv", index=False)

    components = [f"w{idx}" for idx in range(1, 11)] + ["b"]
    roles = FEATURES + ["Intercept"]
    objective_contributions = np.r_[0.5 * theta[:-1] ** 2, 0.0]
    gradient = np.r_[theta[:-1], 0.0]
    hessian_diagonal = np.r_[np.ones(10), 0.0]
    pd.DataFrame(
        {
            "theta_component": components,
            "feature_or_role": roles,
            "optimized_value": theta,
            "objective_contribution_0.5_wj_squared": objective_contributions,
            "first_derivative": gradient,
            "hessian_diagonal": hessian_diagonal,
        }
    ).to_csv(CSV_DIR / "05_theta_gradient_hessian.csv", index=False)

    U = labels[:, None] * np.column_stack([X, np.ones(len(X))])
    signed_margins = U @ theta
    g_values = 1.0 - signed_margins
    support = np.isclose(signed_margins, 1.0, atol=1e-5)
    margin = identifiers.copy()
    for idx, feature in enumerate(FEATURES):
        margin[f"u_{feature}"] = U[:, idx]
    margin["u_b"] = U[:, -1]
    margin["score_s_i"] = scores
    margin["u_i_T_theta"] = signed_margins
    margin["t_i_times_s_i"] = labels * scores
    margin["identity_difference"] = signed_margins - labels * scores
    margin["g_i_equals_1_minus_u_i_T_theta"] = g_values
    margin["constraint_u_i_T_theta_ge_1"] = np.where(g_values <= 1e-8, "Satisfied", "Violated")
    margin["support_vector"] = np.where(support, "Yes", "No")
    margin["predicted_t"] = np.where(scores >= 0, 1, -1)
    margin["correct"] = np.where(margin["predicted_t"].to_numpy() == labels, "Correct", "Wrong")
    margin.to_csv(CSV_DIR / "06_u_margin_check.csv", index=False)

    margin[
        [
            "i",
            "PatientID",
            "class_y_i",
            "t_i",
            "score_s_i",
            "u_i_T_theta",
            "g_i_equals_1_minus_u_i_T_theta",
            "constraint_u_i_T_theta_ge_1",
            "support_vector",
        ]
    ].to_csv(CSV_DIR / "00_main_10row_table.csv", index=False)

    pd.DataFrame(
        [
            ["Parameter vector", "theta = [w1,...,w10,b]^T", "The 11 values optimized in the primal"],
            ["Signed row", "u_i = t_i[x_i1,...,x_i10,1]^T", "One row of U"],
            ["Score", "s_i = w^T x_i + b", "Positive predicts UPA; negative predicts BPA"],
            ["Signed margin", "u_i^T theta = t_i s_i", "A scalar, not a vector"],
            ["Hard-margin rule", "u_i^T theta >= 1", "Equivalent to g_i(theta) <= 0"],
            ["Constraint function", "g_i(theta) = 1-u_i^T theta", "Satisfied when <= 0"],
            ["Objective", "phi(theta) = 0.5||w||^2", "b is excluded"],
            ["First derivative", "grad phi = [w,0]^T", "Shown in 05_theta_gradient_hessian.csv"],
            ["Second derivative", "Hessian phi = diag(I_10,0)", "Positive semidefinite"],
        ],
        columns=["concept", "formula", "interpretation"],
    ).to_csv(CSV_DIR / "README_formulas.csv", index=False)

    pd.DataFrame(
        {
            "metric": [
                "primal_objective_0.5_norm_w_squared",
                "norm_w",
                "boundary_to_support_distance_1_over_norm_w",
                "full_margin_width_2_over_norm_w",
                "minimum_signed_margin",
                "support_vector_count",
                "training_accuracy_teaching_only",
            ],
            "value": [
                0.5 * theta[:-1] @ theta[:-1],
                np.linalg.norm(theta[:-1]),
                1.0 / np.linalg.norm(theta[:-1]),
                2.0 / np.linalg.norm(theta[:-1]),
                signed_margins.min(),
                support.sum(),
                np.mean(np.where(scores >= 0, 1, -1) == labels),
            ],
        }
    ).to_csv(CSV_DIR / "07_optimization_summary.csv", index=False)

    pd.DataFrame(
        [
            ["Full cohort", 114, "-", "No", "No", "No", "N/A"],
            ["CV fold 1", 91, 23, "No", "No", "No", "N/A"],
            ["CV fold 2", 91, 23, "No", "No", "No", "N/A"],
            ["CV fold 3", 91, 23, "No", "No", "No", "N/A"],
            ["CV fold 4", 91, 23, "No", "No", "No", "N/A"],
            ["CV fold 5", 92, 22, "No", "No", "No", "N/A"],
        ],
        columns=[
            "analysis_instance",
            "training_n",
            "held_out_n",
            "hard_margin_feasible",
            "model_fitted",
            "roc_scores_available",
            "auroc",
        ],
    ).to_csv(CSV_DIR / "08_full_data_gate.csv", index=False)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    FIG_DIR.mkdir(parents=True, exist_ok=True)
    frame = pd.read_excel(SOURCE, sheet_name="Sheet1")
    selected = pd.concat(
        [
            frame.loc[frame["Model 1 Target"].eq("UPA")].head(5),
            frame.loc[frame["Model 1 Target"].eq("BPA")].head(5),
        ]
    ).sort_index()
    if len(selected) != 10 or selected["Model 1 Target"].value_counts().to_dict() != {"UPA": 5, "BPA": 5}:
        raise RuntimeError("Expected a balanced ten-row teaching subset.")

    X_raw = selected[FEATURES].apply(pd.to_numeric, errors="coerce")
    preprocessor, X = Preprocessor.fit(X_raw)
    labels = np.where(selected["Model 1 Target"].eq("UPA"), 1.0, -1.0)
    feasible, initial_theta, message = feasibility_point(X, labels)
    if not feasible or initial_theta is None:
        raise RuntimeError(f"Teaching subset is not feasible: {message}")
    theta = fit_primal_hard_margin(X, labels, initial_theta)
    scores = X @ theta[:-1] + theta[-1]
    margins = labels * scores
    if margins.min() < 1 - 1e-6:
        raise RuntimeError(f"Minimum teaching margin is {margins.min():.9f}.")

    wb = Workbook()
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    build_start_sheet(wb)
    build_raw_sheet(wb, selected)
    build_transform_sheet(wb)
    build_standardized_sheet(wb)
    build_theta_sheet(wb, theta)
    build_margin_sheet(wb)
    build_visual_sheet(wb, selected)
    build_full_data_gate(wb)
    base_font(wb)
    wb.active = 0
    wb.save(WORKBOOK_PATH)
    save_score_figure(selected, scores)
    export_csv_tables(selected, X_raw, preprocessor, X, labels, theta, scores)
    print(f"Workbook: {WORKBOOK_PATH}")
    print(f"Figure: {FIGURE_PATH}")
    print(f"CSV tables: {CSV_DIR}")
    print(f"Minimum signed margin: {margins.min():.9f}")
    print(f"Objective: {0.5 * theta[:-1] @ theta[:-1]:.9f}")


if __name__ == "__main__":
    main()
